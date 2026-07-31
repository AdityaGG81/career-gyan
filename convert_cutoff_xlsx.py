"""
Convert Maharashtra MHT-CET Engineering Cutoffs Excel to JSON + CSV.
Handles the actual Excel structure with multiple sheets and header rows.
"""
import json
import sys
import os
import csv
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, '..', 'Maharashtra_MHT_CET_Engineering_Cutoffs_2025 (1).xlsx')
JSON_OUTPUT = os.path.join(SCRIPT_DIR, 'database', 'data', 'mht_cet_cutoffs.json')
CSV_OUTPUT = os.path.join(SCRIPT_DIR, 'public', 'downloads', 'Maharashtra_MHT_CET_Engineering_Cutoffs_2025.csv')

print(f"Reading Excel: {EXCEL_PATH}")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

records = []

# Process both data sheets
for sheet_name in ['Popular Colleges', 'All OPEN Cutoffs']:
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found, skipping.")
        continue
    
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    # Find the header row (contains 'Institute Code')
    header_idx = None
    for i, row in enumerate(rows):
        if row and any(str(cell).strip() == 'Institute Code' for cell in row if cell):
            header_idx = i
            break
    
    if header_idx is None:
        print(f"Sheet '{sheet_name}': Could not find header row, skipping.")
        continue
    
    header = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[header_idx])]
    print(f"Sheet '{sheet_name}': header at row {header_idx}, columns: {header}")
    
    count = 0
    for row in rows[header_idx + 1:]:
        if not row or all(cell is None for cell in row):
            continue
        
        record = {}
        for i, val in enumerate(row):
            if i < len(header):
                record[header[i]] = val
        
        # Extract fields
        institute_code = record.get('Institute Code')
        institute_name = record.get('Institute Name', '')
        course_code = record.get('Course Code')
        branch = record.get('Branch / Course', '')
        seat_type = record.get('Seat Type', '')
        merit_no = record.get('Merit No.')
        percentile = record.get('MHT-CET Percentile', 0)
        percentile_band = record.get('Percentile Band', '')
        
        if not institute_name or institute_name == 'Institute Name':
            continue
        
        try:
            percentile = float(percentile) if percentile is not None else 0.0
        except (ValueError, TypeError):
            percentile = 0.0
        
        try:
            merit_no = int(merit_no) if merit_no is not None else None
        except (ValueError, TypeError):
            merit_no = None
        
        try:
            institute_code = int(institute_code) if institute_code is not None else None
        except (ValueError, TypeError):
            institute_code = None
        
        normalized = {
            'college_code': institute_code,
            'college_name': str(institute_name).strip(),
            'branch_code': str(course_code).strip() if course_code else None,
            'branch_name': str(branch).strip(),
            'category': str(seat_type).strip() if seat_type else 'GOPENS',
            'category_full': 'General OPEN - State Level' if seat_type == 'GOPENS' else ('General OPEN - Home University' if seat_type == 'GOPENH' else str(seat_type)),
            'percentile': percentile,
            'year': 2025,
            'round': 'CAP Round I',
            'status': None,
            'quota': 'MH',
            'merit_no': merit_no,
            'percentile_band': str(percentile_band).strip() if percentile_band else None,
        }
        
        records.append(normalized)
        count += 1
    
    print(f"  -> Extracted {count} records from '{sheet_name}'")

wb.close()

# Deduplicate (Popular Colleges data may overlap with All OPEN Cutoffs)
seen = set()
unique_records = []
for r in records:
    key = (r['college_code'], r['branch_code'], r['category'])
    if key not in seen:
        seen.add(key)
        unique_records.append(r)

print(f"\nTotal unique records: {len(unique_records)} (from {len(records)} raw)")

# Stats
colleges = set(r['college_name'] for r in unique_records)
branches = set(r['branch_name'] for r in unique_records)
categories = set(r['category'] for r in unique_records)

print(f"Unique colleges: {len(colleges)}")
print(f"Unique branches: {len(branches)}")
print(f"Categories: {sorted(categories)}")
print(f"Sample: {json.dumps(unique_records[0], default=str, indent=2)}")

# Save JSON
os.makedirs(os.path.dirname(JSON_OUTPUT), exist_ok=True)
with open(JSON_OUTPUT, 'w', encoding='utf-8') as f:
    json.dump(unique_records, f, ensure_ascii=False, default=str)
print(f"\nJSON saved: {JSON_OUTPUT} ({os.path.getsize(JSON_OUTPUT):,} bytes)")

# Save CSV for download
os.makedirs(os.path.dirname(CSV_OUTPUT), exist_ok=True)
csv_headers = ['Institute Code', 'Institute Name', 'Course Code', 'Branch / Course', 'Seat Type', 'Merit No.', 'MHT-CET Percentile', 'Percentile Band', 'Round', 'Year']
with open(CSV_OUTPUT, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(csv_headers)
    for r in unique_records:
        writer.writerow([
            r.get('college_code', ''),
            r.get('college_name', ''),
            r.get('branch_code', ''),
            r.get('branch_name', ''),
            r.get('category', ''),
            r.get('merit_no', ''),
            r.get('percentile', ''),
            r.get('percentile_band', ''),
            r.get('round', 'CAP Round I'),
            r.get('year', 2025),
        ])
print(f"CSV saved: {CSV_OUTPUT} ({os.path.getsize(CSV_OUTPUT):,} bytes)")
print("\nDone!")
